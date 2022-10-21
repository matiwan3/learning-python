from faker import Faker
from openpyxl import Workbook
import pandas as pd

wb = Workbook()
ws = wb.active
fake_data = Faker(['pl_PL'])

for i in range(1,110):
    for j in range(1,4):
        ws.cell(row=i, column=1). value = fake_data.name()
        ws.cell(row=i, column=2). value = fake_data.email()
        ws.cell(row=i, column=3). value = fake_data.address()
        ws.cell(row=i, column=4). value = fake_data.job()
wb.save("testData.xlsx")